import os
import pandas as pd
import numpy as np
from typing import List, Dict, Any, Union, Optional

class DataWrangler:
    """Helper class for cleaning, merging, and exporting Excel/CSV data in Xenia."""

    @classmethod
    def clean_dataset(cls, data: Union[str, pd.DataFrame, List[Dict[str, Any]]], 
                      drop_empty_rows: bool = True, 
                      drop_duplicates: bool = True,
                      fill_na_value: Optional[Any] = None) -> pd.DataFrame:
        """Cleans a dataset, dropping duplicates and empty rows, and returning a DataFrame."""
        df = cls._to_dataframe(data)
        
        if df.empty:
            return df
            
        if drop_empty_rows:
            # Drop rows where all columns are NaN/empty
            df = df.dropna(how='all')
            
        if drop_duplicates:
            df = df.drop_duplicates()
            
        if fill_na_value is not None:
            df = df.fillna(fill_na_value)
            
        # Reset index after cleaning
        df = df.reset_index(drop=True)
        return df

    @classmethod
    def merge_datasets(cls, left: Union[str, pd.DataFrame, List[Dict[str, Any]]], 
                       right: Union[str, pd.DataFrame, List[Dict[str, Any]]], 
                       on_column: str, 
                       join_type: str = "inner") -> pd.DataFrame:
        """Merges two datasets on a specific column using pandas merge (join_type: inner, left, right, outer)."""
        df_left = cls._to_dataframe(left)
        df_right = cls._to_dataframe(right)
        
        if df_left.empty:
            return df_right
        if df_right.empty:
            return df_left
            
        # Ensure the column exists in both
        if on_column not in df_left.columns:
            # Case insensitive check
            cols_l = {c.lower(): c for c in df_left.columns}
            if on_column.lower() in cols_l:
                df_left = df_left.rename(columns={cols_l[on_column.lower()]: on_column})
            else:
                raise ValueError(f"Column '{on_column}' not found in left dataset columns: {list(df_left.columns)}")
                
        if on_column not in df_right.columns:
            cols_r = {c.lower(): c for c in df_right.columns}
            if on_column.lower() in cols_r:
                df_right = df_right.rename(columns={cols_r[on_column.lower()]: on_column})
            else:
                raise ValueError(f"Column '{on_column}' not found in right dataset columns: {list(df_right.columns)}")

        # Clean key columns to string, stripped, to prevent merge failures due to leading/trailing spaces
        df_left[on_column] = df_left[on_column].astype(str).str.strip()
        df_right[on_column] = df_right[on_column].astype(str).str.strip()

        merged_df = pd.merge(df_left, df_right, on=on_column, how=join_type)
        return merged_df

    @classmethod
    def detect_anomalies(cls, data: Union[str, pd.DataFrame, List[Dict[str, Any]]], 
                         column: str, 
                         threshold_std: float = 3.0) -> pd.DataFrame:
        """Finds rows containing values in a numeric column that are more than threshold_std standard deviations from the mean."""
        df = cls._to_dataframe(data)
        if df.empty or column not in df.columns:
            return pd.DataFrame()
            
        # Ensure column is numeric
        numeric_series = pd.to_numeric(df[column], errors='coerce')
        valid_data = df[numeric_series.notna()]
        
        if valid_data.empty:
            return pd.DataFrame()
            
        mean = numeric_series.mean()
        std = numeric_series.std()
        
        if std == 0:
            return pd.DataFrame()
            
        z_scores = (numeric_series - mean) / std
        anomalies = df[z_scores.abs() > threshold_std]
        return anomalies

    @classmethod
    def export_to_excel(cls, data: Union[pd.DataFrame, List[Dict[str, Any]]], 
                        file_path: str, 
                        sheet_name: str = "Sheet1", 
                        styled: bool = True) -> str:
        """Saves a dataset to a styled Excel worksheet. Adjusts column widths automatically and applies corporate styles."""
        df = cls._to_dataframe(data)
        abs_path = os.path.abspath(file_path)
        os.makedirs(os.path.dirname(abs_path), exist_ok=True)
        
        if not styled:
            df.to_excel(abs_path, sheet_name=sheet_name, index=False)
            return abs_path

        # Write styled output using openpyxl
        with pd.ExcelWriter(abs_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Apply styling
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            
            # Styling definitions
            header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid") # Corporate Blue
            cell_font = Font(name="Segoe UI", size=10)
            align_center = Alignment(horizontal="center", vertical="center")
            align_left = Alignment(horizontal="left", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color='D3D3D3'),
                right=Side(style='thin', color='D3D3D3'),
                top=Side(style='thin', color='D3D3D3'),
                bottom=Side(style='thin', color='D3D3D3')
            )
            
            # Format Headers
            for col_idx in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col_idx)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = align_center
                cell.border = thin_border
                
            # Format Cells and Autofit Column Widths
            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    # Apply general styling to all data rows
                    if cell.row > 1:
                        cell.font = cell_font
                        cell.border = thin_border
                        # Align numbers to right, strings to left
                        if isinstance(cell.value, (int, float)):
                            cell.alignment = Alignment(horizontal="right", vertical="center")
                        else:
                            cell.alignment = align_left
                            
                    # Calculate max length
                    val_str = str(cell.value or '')
                    if len(val_str) > max_len:
                        max_len = len(val_str)
                        
                # Set dynamic column width with buffer padding
                worksheet.column_dimensions[col_letter].width = max(max_len + 3, 12)
                
        return abs_path

    @classmethod
    def _to_dataframe(cls, data: Union[str, pd.DataFrame, List[Dict[str, Any]]]) -> pd.DataFrame:
        """Converts diverse data types into a pandas DataFrame."""
        if isinstance(data, pd.DataFrame):
            return data.copy()
            
        if isinstance(data, list):
            if not data:
                return pd.DataFrame()
            return pd.DataFrame(data)
            
        if isinstance(data, str):
            # Resolve path and check extension
            if not os.path.exists(data):
                raise FileNotFoundError(f"File not found for wrangler: {data}")
            ext = os.path.splitext(data)[1].lower()
            if ext == '.csv':
                return pd.read_csv(data)
            elif ext in ('.xls', '.xlsx'):
                return pd.read_excel(data)
            else:
                raise ValueError(f"Unsupported file format: {ext}")
                
        raise TypeError(f"Invalid input type for DataWrangler: {type(data)}")
